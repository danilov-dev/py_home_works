import json
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


class Currency:
    def __init__(self, digit_code:str, char_code:str, count:int, name:str, rate:float):
        self.digit_code = digit_code
        self.char_code = char_code
        self.count = count
        self.name = name
        self.rate = rate

    def to_dict(self):
        return {
            "digit_code": self.digit_code,
            "char_code": self.char_code,
            "count": self.count,
            "name": self.name,
            "rate": self.rate
        }

class CurrencyParser:
    def __init__(self, url:str) -> None:
        self.url = url
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36"
        }
        self._soup = None
        self.currencies = []

    def _get_soup(self):
        response = requests.get(self.url, headers=self.headers, timeout=15)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            self._soup = soup
        else:
            raise ConnectionError

    def _parse_data(self):
        table = self._soup.find("table", class_="data")
        if table is None:
            raise Exception("Данных для парсинга не найдено")
        rows = table.find_all("tr")
        for row in rows[1:]:
            tds = row.find_all("td")
            if len(tds)<5:
                continue
            try:
                currency = Currency(
                    digit_code=tds[0].text.strip(),
                    char_code=tds[1].text.strip(),
                    count=int(tds[2].text.strip()),
                    name=tds[3].text.strip(),
                    rate=float(tds[4].text.replace(",", ".").strip()),
                )
                self.currencies.append(currency)
            except Exception as e:
                print(e)


    def parse(self):
        self._get_soup()
        self._parse_data()
        return self.currencies


class FileSaver:

    def save_to_json(self, file_path: str, data: dict) -> None:
        with open(file_path+'.json', 'w', encoding='utf-8') as f:
            try:
                output = {
                    "currencies": [currency.to_dict() for currency in data['currencies']]
                }
                json_string = json.dumps(output, ensure_ascii=False, indent=2)
                f.write(json_string)
                print("JSON записан")
            except Exception as e:
                print(f"Error: {e}")

    def save_to_excel(self, file_path: str, data: dict) -> None:
        try:
            currencies = data.get("currencies", [])
            if not currencies:
                print("Нет данных для сохранения в Excel")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Курсы валют"

            headers = ["digit_code", "char_code", "count", "name", "rate"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row_idx, currency in enumerate(currencies, start=2):
                ws.cell(row=row_idx, column=1, value=currency.digit_code)
                ws.cell(row=row_idx, column=2, value=currency.char_code)
                ws.cell(row=row_idx, column=3, value=currency.count)
                ws.cell(row=row_idx, column=4, value=currency.name)
                ws.cell(row=row_idx, column=5, value=currency.rate)

            wb.save(f"{file_path}.xlsx")
            print(f"Excel файл успешно создан: {file_path}.xlsx")
        except Exception as e:
            print(f"Ошибка при сохранении Excel: {e}")


class ParserManager:
    def __init__(self, url:str):
        self.url = url
        self.parsers = CurrencyParser(url)
        self.data = {}

    def parse(self):
        self.data["currencies"] = self.parsers.parse()
        if self.data["currencies"]:
            print("Данные удачно собраны")

    def get_data(self):
        if self.data:
            return self.data
        else:
            return None

    def save_to_file(self, file_name:str, selection_num: int) -> None:
        saver = FileSaver()
        match selection_num:
            case 1:
                saver.save_to_json(file_name, self.data)
            case 2:
                saver.save_to_excel(file_name, self.data)

def choose_extension():
    while True:
        print("Как сохранить данные?")
        print("1. JSON")
        print("2. Excel")
        file_extension = input(">>> ")
        if file_extension.isdigit():
            return int(file_extension)

if __name__ == "__main__":
    print("Начнем парсинг!")
    base_url = "https://www.cbr.ru/currency_base/daily/"
    try:
        manager = ParserManager(base_url)
        manager.parse()

        request = choose_extension()

        manager.save_to_file("data",request)
    except Exception as e:
        print(e)






