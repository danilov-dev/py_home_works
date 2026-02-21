import json
import time
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
from requests.exceptions import RequestException


def _get_soup(url, headers):
    """Получение HTML и создание объекта BeautifulSoup с задержкой"""
    try:
        print("Ожидание 2 секунды перед запросом...")
        time.sleep(2)

        response = requests.get(url, headers=headers, timeout=15)
        response.encoding = 'utf-8'

        if response.status_code != 200:
            print(f"Ошибка HTTP: {response.status_code}")
            return None

        soup = BeautifulSoup(response.text, 'html.parser')
        return soup

    except RequestException as e:
        print(f"Ошибка при выполнении запроса: {e}")
        return None
    except Exception as e:
        print(f"Неожиданная ошибка: {e}")
        return None


def _get_headers(soup):
    """Извлечение заголовков таблицы"""
    headers = []
    first_row = soup.find_all('th')
    if first_row:
        headers = [header.text.strip() for header in first_row]
    return headers


def parse_data(soup):
    """Парсинг данных о валютах"""
    data = []

    currency_table = soup.find('table', class_='data')
    if currency_table is None:
        print("Таблица с данными не найдена")
        return []

    rows = currency_table.find_all('tr')

    for row in rows[1:]:
        tds = row.find_all('td')

        if len(tds) < 5:
            continue

        try:
            currency = {
                "digit_code": tds[0].text.strip(),
                "char_code": tds[1].text.strip(),
                "count": tds[2].text.strip(),
                "name": tds[3].text.strip(),
                "rate": tds[4].text.strip().replace(',', '.')  # Заменяем запятую на точку для числового формата
            }
            data.append(currency)
        except (IndexError, AttributeError) as e:
            print(f"Ошибка при обработке строки: {e}")
            continue

    print(f"Найдено {len(data)} валют")
    return data


def _save_json(data):
    """Сохранение данных в JSON файл"""
    try:
        with open('currency.json', 'w', encoding='utf-8') as f:
            json_string = json.dumps(data, ensure_ascii=False, indent=2)
            f.write(json_string)
            print("Данные успешно сохранены в файл currency.json")
    except IOError as e:
        print(f"Ошибка при записи JSON файла: {e}")
    except Exception as e:
        print(f"Неожиданная ошибка при сохранении JSON: {e}")


def _save_excel(data, headers):
    """Сохранение данных в Excel файл"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Курсы валют"

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row, currency in enumerate(data, 2):
            ws.cell(row=row, column=1,
                    value=int(currency["digit_code"]) if currency["digit_code"].isdigit() else currency["digit_code"])
            ws.cell(row=row, column=2, value=currency["char_code"])
            ws.cell(row=row, column=3, value=int(currency["count"]))
            ws.cell(row=row, column=4, value=currency["name"])

            rate_value = float(currency["rate"].replace(',', '.'))
            ws.cell(row=row, column=5, value=rate_value)

        wb.save('currency.xlsx')
        print("Данные успешно сохранены в файл currency.xlsx")

    except IOError as e:
        print(f"Ошибка при записи Excel файла: {e}")
    except ValueError as e:
        print(f"Ошибка преобразования данных: {e}")
    except Exception as e:
        print(f"Неожиданная ошибка при сохранении Excel: {e}")


def get_currency_data(url):
    """Основная функция для получения данных о курсах валют"""
    try:
        requests.get("https://www.google.com", timeout=5)
    except RequestException:
        print("Нет подключения к интернету. Проверьте соединение.")
        return

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7"
    }

    print("Начинаем парсинг данных с сайта ЦБ РФ...")
    soup = _get_soup(url, headers)

    if soup is None:
        print("Не удалось получить данные с сайта")
        return

    table_headers = _get_headers(soup)
    if not table_headers:
        print("Не удалось получить заголовки таблицы")
        return

    data = parse_data(soup)

    if not data:
        print("Нет данных для сохранения")
        return

    while True:
        print("\nВыберите формат сохранения данных:")
        print("1. JSON")
        print("2. Excel")
        print("3. Оба формата")
        print("4. Выход")

        choice = input(">>> ").strip()

        if choice == '1':
            _save_json(data)
            break
        elif choice == '2':
            _save_excel(data, table_headers)
            break
        elif choice == '3':
            _save_json(data)
            _save_excel(data, table_headers)
            break
        elif choice == '4':
            print("Программа завершена")
            break
        else:
            print("Неверный выбор. Пожалуйста, выберите 1, 2, 3 или 4")


def main():
    """Точка входа в программу"""
    url = "https://www.cbr.ru/currency_base/daily/"

    try:
        get_currency_data(url)
    except KeyboardInterrupt:
        print("\nПрограмма прервана пользователем")
    except Exception as e:
        print(f"Критическая ошибка: {e}")
    finally:
        print("Работа программы завершена")


if __name__ == "__main__":
    main()